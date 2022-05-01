from openpyxl import load_workbook

from . import classes
from . import file_operations

def get_room_mapping_from_excel_file(filepath):
    # a função carrega dados de uma planilha; após isso, cria e retorna um dicionário da forma
    # {day: {hour: {classroom: [component_code, component_class] ...} ...} ...}

    excel_workbook = load_workbook(filepath)
    worksheet = excel_workbook.active
    room_mapping = {}

    for i in range(2, worksheet.max_row + 1):
        day = worksheet.cell(row = i, column = 1).value
        hour = worksheet.cell(row = i, column = 2).value
        classroom = worksheet.cell(row = i, column = 3).value
        component_code = worksheet.cell(row = i, column = 4).value
        component_class = worksheet.cell(row = i, column = 5).value

        if day not in room_mapping:
            room_mapping[day] = {}
            room_mapping[day][hour] = {}
        elif hour not in room_mapping[day]:
            room_mapping[day][hour] = {}
        
        if classroom not in room_mapping[day][hour]:
            room_mapping[day][hour][classroom] = []
        
        room_mapping[day][hour][classroom] = [component_code, component_class]

    return room_mapping

def get_academics_mapping_from_pickle_file(filepath):
    data = file_operations.load_pickle_file(filepath)

    academics_mapping = {}

    students = []
    professors = []

    professors_ids = []
    students_ids = []

    for infos in data['data']:
        infos_subject = infos[0] # duas colunas, contendo as informações da disciplina; primeira com os nomes das infos e segunda com: disciplina, turma, nome do docente
        infos_subject = infos_subject[1] # a segunda coluna possui os valores que estamos querendo utilizar

        component_code = infos_subject[0] # a primeira linha tem o código e o nome da disciplina
        component_code = component_code.split('-')[0].strip() # "código - nome" -> "código " -> "código"

        component_class = infos_subject[1].strip() # segunda linha tem o código da turma

        professor_name = infos_subject[2] # a terceira linha tem o nome do professor

        infos_students = infos[1] # informações dos alunos: matrícula, nome, curso

        students_registration_numbers = infos_students['Matrícula'].tolist() # lista com as matrículas
        students_courses = infos_students['Curso'].tolist() # lista com os cursos

        if component_code not in academics_mapping: # adiciona o código da disciplina apenas quando aparece pela primeira vez
            academics_mapping[component_code] = {}
        
        if component_class not in academics_mapping[component_code]: # os estudantes em uma turma de um determinado código de disciplina são os mesmos, então só adicionamos na primeira vez que aparecem
            academics_mapping[component_code][component_class] = []

            if professor_name not in professors_ids:
                component_professor = classes.Professor(name = professor_name)
                professors_ids.append(professor_name)
                professors.append(component_professor)
            else:
                index = professors_ids.index(professor_name)
                component_professor = professors[index]

            academics_mapping[component_code][component_class].append(component_professor)

            for i in range(0, len(students_registration_numbers)):
                student_registration_number = str(students_registration_numbers[i])
                student_course = students_courses[i]
                student_course = student_course.replace("- NOVA IGUAÇU", "").split("- Presencial")[0].split("-")
                student_course = " - ".join([string.strip() for string in student_course])

                if student_registration_number not in students_ids:
                    student = classes.Student(registration_number = student_registration_number, course = student_course)
                    students_ids.append(student_registration_number)
                    students.append(student)
                else:
                    index = students_ids.index(student_registration_number)
                    student = students[index]
    
                academics_mapping[component_code][component_class].append(student)

    return academics_mapping, students, professors

def remove_not_imported_component_code(room_mapping, academics_mapping):
    remove = []
    for day in room_mapping:
        for hour in room_mapping[day]:
            for classroom in room_mapping[day][hour]:
                component_code, _ = room_mapping[day][hour][classroom]
                if component_code not in academics_mapping:
                    remove.append([day, hour, classroom])
    
    for day, hour, classroom in remove:                
        del room_mapping[day][hour][classroom]

def remove_not_imported_component_class(room_mapping, academics_mapping):
    remove = []
    for day in room_mapping:
        for hour in room_mapping[day]:
            for classroom in room_mapping[day][hour]:
                component_code, component_class = room_mapping[day][hour][classroom]
                if component_class not in academics_mapping[component_code]:
                    remove.append([day, hour, classroom])

    for day, hour, classroom in remove:                
        del room_mapping[day][hour][classroom]

def remove_duplicated_component_class(room_mapping, academics_mapping, classrooms):
    components = {}

    for day in room_mapping:
        for hour in room_mapping[day]:
            for classroom in room_mapping[day][hour]:
                component_code, component_class = room_mapping[day][hour][classroom]
                component = component_code + " - " + component_class

                current_classroom = find_class_instance_by_id(classroom, classrooms)

                if component not in components:
                    components[component] = {}
                
                if classroom not in components[component]:
                    components[component][classroom] = []
                
                components[component][classroom].append([day, hour, current_classroom.type.type, len(academics_mapping[component_code][component_class])])

    priority_list = {
        45: ["Sala grande", "Sala pós", "Sala média", "Sala pequena"], # a partir de 45 alunos
        25: ["Sala pós", "Sala média", "Sala grande", "Sala pequena"], # entre 25 e 44 alunos
        0: ["Sala pequena", "Sala média", "Sala pós", "Sala grande"] # entre 0 e 24 alunos
    }

    for component in components:
        if len(components[component]) > 1:
            fitted_option = [None, 3]
            for classroom in components[component]:
                option = components[component][classroom][0]
                count_students = list(priority_list.keys())

                for count in count_students:
                    if option[3] >= count:
                        priority = priority_list[count].index(option[2])

                        if priority <= fitted_option[1]:
                            fitted_option = [classroom, priority]
                        break
                    
            for classroom in components[component]:
                if fitted_option[0] != classroom:
                    for classroom_list in components[component][classroom]:
                        day = classroom_list[0]
                        hour = classroom_list[1]
                        del room_mapping[day][hour][classroom]

def fill_empty_component_class(room_mapping, academics_mapping):
    classes_to_fix = {}

    for component_code in academics_mapping:
        classes_to_fix[component_code] = list(academics_mapping[component_code].keys()) # pega todas as turmas de uma disciplina importadas do sigaa

    for _ in range(0, 3): # normalmente uma disciplina tem duas turmas, então fazemos três iterações para garantir que as modificações sejam completas
        values_to_remove = []
        for day in room_mapping:
            for hour in room_mapping[day]:
                for classroom in room_mapping[day][hour]:
                    component_code, component_class = room_mapping[day][hour][classroom]
                    classes_count = len(classes_to_fix[component_code]) # quantidade de turmas de uma disciplina
                    if component_class == None:
                        if classes_count == 1:
                            # se não tem turma especificada no pdf, mas tem no sigaa, usa a do sigaa
                            room_mapping[day][hour][classroom][1] = classes_to_fix[component_code][0]
                            if (component_code, classes_to_fix[component_code][0]) not in values_to_remove:
                                values_to_remove.append((component_code, classes_to_fix[component_code][0])) # anota que essa turma da disciplina já está correta
                    else:
                        if component_class in classes_to_fix[component_code]:
                            if (component_code, component_class) not in values_to_remove:
                                values_to_remove.append((component_code, component_class)) # anota que essa turma da disciplina já está correta

        for component_code, component_class in values_to_remove:
            classes_to_fix[component_code].remove(component_class) # remove as classes corrigidas

def get_classroom_types_from_excel_file(filepath):
    excel_workbook = load_workbook(filepath)
    worksheet = excel_workbook.active
    classroom_types = []

    for i in range(2, worksheet.max_row + 1):
        type = worksheet.cell(row = i, column = 1).value
        height = worksheet.cell(row = i, column = 2).value
        width = worksheet.cell(row = i, column = 3).value
        length = worksheet.cell(row = i, column = 4).value

        classroom_types.append(classes.Classroom_Type(type, float(height), float(width), float(length)))

    return classroom_types

def get_classrooms_from_excel_file(filepath, classroom_types):
    excel_workbook = load_workbook(filepath)
    worksheet = excel_workbook.active
    classrooms = []

    for i in range(2, worksheet.max_row + 1):
        id = worksheet.cell(row = i, column = 1).value
        type = worksheet.cell(row = i, column = 2).value
        
        for classroom_type in classroom_types:
            if classroom_type.type == type:
                break

        classroom = classes.Classroom(id, classroom_type)
        classrooms.append(classroom)

    return classrooms

def get_academics_dict(professors, students):
    academics = {}
    id_count = 1

    for professor in professors:
        id = "ACADÊMICO_" + str(id_count)
        academics[id] = (professor.id, "")
        professor.academic_id = id
        id_count += 1

    for student in students:
        id = "ACADÊMICO_" + str(id_count)
        academics[id] = ("", student.id)
        student.academic_id = id
        id_count += 1
    
    return academics

def get_processed_data():
    filename = "lib/file.txt"
    parameters = file_operations.load_parameters_from_txt_file(filename)
    pickle_filepath, room_mapping_filepath, classroom_types_filepath, classrooms_filepath = parameters
    
    room_mapping = get_room_mapping_from_excel_file(room_mapping_filepath)
    academics_mapping, students, professors = get_academics_mapping_from_pickle_file(pickle_filepath)

    classroom_types = get_classroom_types_from_excel_file(classroom_types_filepath)
    classrooms = get_classrooms_from_excel_file(classrooms_filepath, classroom_types)

    remove_not_imported_component_code(room_mapping, academics_mapping)
    fill_empty_component_class(room_mapping, academics_mapping)
    remove_not_imported_component_class(room_mapping, academics_mapping)
    remove_duplicated_component_class(room_mapping, academics_mapping, classrooms)

    academics = anonymize_data(room_mapping, academics_mapping, professors, students)

    return room_mapping, academics_mapping, academics, professors, students, classroom_types, classrooms

def anonymize_dict(dict, prefix):
    id_count = 1
    dict_keys = list(dict.keys())
    while dict_keys != []:
        new_key = prefix + "_" + str(id_count)
        dict[new_key] = dict.pop(dict_keys.pop(0))

        id_count += 1

def anonymize_list(list, prefix):
    id_count = 1

    for element in list:
        id = prefix + "_" + str(id_count)
        element.id = id
        id_count += 1

def anonymize_data(room_mapping, academics_mapping, professors, students):
    anonymize_list(professors, "PROFESSOR")
    anonymize_list(students, "ALUNO")
    academics = get_academics_dict(professors, students)

    # dicionário com key = old_key; value = old_key
    component_codes = {}
    
    for component_code in list(academics_mapping.keys()):
        component_codes[component_code] = component_code

    # dicionário com key = new_key; value = old_key
    anonymize_dict(component_codes, "DISCIPLINA")
    anonymize_dict(academics_mapping, "DISCIPLINA")

    for day in room_mapping:
        for hour in room_mapping[day]:
            for classroom in room_mapping[day][hour]:
                old_component_code, _ = room_mapping[day][hour][classroom]
                for new_component_code in component_codes:
                    if component_codes[new_component_code] == old_component_code:
                        room_mapping[day][hour][classroom][0] = new_component_code
                        break

    return academics

def find_class_instance_by_id(id, list):
    for element in list:
        if id == element.id:
            return element
    
    return None

def find_class_instance_by_academic_id(academic_id, academics, professors, students):
    professor_id, student_id = academics[academic_id]

    if professor_id == "":
        academic = find_class_instance_by_id(student_id, students)

    elif student_id == "":
        academic = find_class_instance_by_id(professor_id, professors)

    return academic

def count_students_by_course(data):
    _, _, _, _, students, _, _ = data

    students_by_courses = {}

    for student in students:
        student_course = student.course
        student_course = student_course.replace("- LICENCIATURA", "(L)").replace("- BACHARELADO", "(B)")

        if student_course not in students_by_courses:
            students_by_courses[student_course] = 0
        
        students_by_courses[student_course] += 1

    with open("count_students_by_course.csv", "w") as file:
        file.write("Curso;Quantidade de alunos\n")
        for item in sorted(students_by_courses.items()):
            file.write("%s; %d\n" % (item[0], item[1]))
    
    return students_by_courses

def count_classroom_types(data):
    room_mapping, _, _, _, _, _, classrooms = data

    classroom_types = {}
    for day in room_mapping:
        for hour in room_mapping[day]:
            for classroom in room_mapping[day][hour]:
                current_classroom = find_class_instance_by_id(classroom, classrooms)
                if current_classroom.type.type not in classroom_types:
                    classroom_types[current_classroom.type.type] = []
                    classroom_types[current_classroom.type.type].append(classroom)
                
                elif classroom not in classroom_types[current_classroom.type.type]:
                    classroom_types[current_classroom.type.type].append(classroom)

    with open("count_classroom_types.csv", "w") as file:
        file.write("Tipo de sala;Quantidade\n")
        for classroom_type in classroom_types:
            file.write("%s;%d\n" % (classroom_type, len(classroom_types[classroom_type])))

def get_classes_predominance(data, courses):
    room_mapping, academics_mapping, _, _, _, _, classrooms = data
    
    courses.sort()
    classes_predominance = {}

    for day in room_mapping:
        for hour in room_mapping[day]:
            for classroom in room_mapping[day][hour]:
                current_classroom = find_class_instance_by_id(classroom, classrooms)
                component_code, component_class = room_mapping[day][hour][classroom]
                component = component_code + " - " + component_class + " - " + current_classroom.type.type

                print(component_code, ";", component_class, ";", classroom, ";", current_classroom.type.type)
                if component not in classes_predominance:
                    classes_predominance[component] = {}
                    for course in courses:
                        classes_predominance[component][course] = 0

                    academics_list = academics_mapping[component_code][component_class]
                    for academic in academics_list:
                        if type(academic) == classes.Student:
                            student_course = academic.course.replace("- LICENCIATURA", "(L)").replace("- BACHARELADO", "(B)")
                            classes_predominance[component][student_course] += 1

    with open("classes_predominance.csv", "w") as file:
        file.write("Disciplina - Turma - Tipo de sala")
        
        for course in courses:
            file.write(";%s" % course)

        file.write("\n")

        for component in classes_predominance:
            file.write(component)
            
            for course in courses:
                file.write(";%s" % classes_predominance[component][course])

            file.write("\n")
        