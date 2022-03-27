from openpyxl import load_workbook

def merge_worksheets(workbook, worksheet1_name, worksheet2_name):
    worksheet1 = workbook[worksheet1_name]
    worksheet2 = workbook[worksheet2_name]

    row_start = worksheet1.max_row + 1 # a iteração vai começar após a última linha preenchida de worksheet1
    row_end = worksheet1.max_row + worksheet2.max_row # a iteração vai terminar quando worksheet1 tiver todas as linhas de worksheet2
    column_start = 1
    column_end = worksheet1.max_column + 1

    row_count = 2

    for i in range(row_start, row_end):
        for j in range(column_start, column_end):
            worksheet1.cell(row = i, column = j).value = worksheet2.cell(row = row_count, column = j).value

        row_count += 1

def updating_worksheets(workbook):
    workbook['Manhã'].title = 'Horários'

    workbook.remove(workbook['Tarde'])
    workbook.remove(workbook['Noite'])

def removing_cells(workbook):
    worksheet = workbook['Horários']
    current_row = 2
    fixed_column = 4

    while(current_row != worksheet.max_row):
        current_cell = worksheet.cell(row = current_row, column = fixed_column)
        if (current_cell.value is None) or (current_cell.value == ' '):
            worksheet.delete_rows(current_row)
        else:
            current_row += 1
        
def save_merged_worksheet(original_filepath, new_filepath):
    excel_workbook = load_workbook(original_filepath)

    merge_worksheets(excel_workbook, 'Manhã', 'Tarde')
    merge_worksheets(excel_workbook, 'Manhã', 'Noite')

    updating_worksheets(excel_workbook)
    removing_cells(excel_workbook)

    excel_workbook.save(new_filepath)

def main():
    save_merged_worksheet("arquivo_turnos_separados.xlsx", "arquivo_turnos_unificados.xlsx")

if __name__ == "__main__":
    main()