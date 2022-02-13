import os
import pickle
import random
from openpyxl import load_workbook
from optparse import OptionParser

class Student:
  def __init__(self, id, course, hasCovid = False, canTransmit = False, hasSymptoms = False, countInfectedTime = 0, infectionProbability = 0.0):
    self.id = id
    self.course = course
    self.hasCovid = hasCovid
    self.canTransmit = canTransmit
    self.hasSymptoms = hasSymptoms
    self.countInfectedTime = countInfectedTime
    self.infectionProbability = infectionProbability

def parse_terminal_options():
    # a função gera padrões de parâmetros a serem analisados e retorna o resultado final

    parser = OptionParser()
    parser.add_option("-d", "--datadir", type="string", dest="datadir", default=".", help="Diretório para salvar dados")
    parser.add_option("-v", "--verbose", action="store_true", dest="verbose", default=False, help="Imprime mensagens para debug")

    return parser.parse_args()

def load_pickle_file(dfpath):
    if os.path.exists(dfpath):
        print("Lendo arquivo pickle.")
        file = open(dfpath, 'rb')
        data = pickle.load(file)
        file.close()
        # print(data)
        return data
    else:
        print("Arquivo não encontrado.")

def get_room_mapping_from_excel_file(filepath, worksheet_name):
    # a função carrega dados de uma planilha; após isso, cria e retorna um dicionário da forma {day: {hour: [room, component_code, component_class] ...} ...}

    excel_workbook = load_workbook(filepath)
    worksheet = excel_workbook[worksheet_name]
    room_mapping = {}

    # code_class_mapping = {}

    for i in range(2, worksheet.max_row + 1):
        day = worksheet.cell(row = i, column = 1).value
        hour = worksheet.cell(row = i, column = 2).value
        classroom = worksheet.cell(row = i, column = 3).value
        component_code = worksheet.cell(row = i, column = 4).value
        component_class = worksheet.cell(row = i, column = 5).value

        if day not in room_mapping:
            # print(day)
            # print(hour)
            room_mapping[day] = {}
            room_mapping[day][hour] = {}
        elif hour not in room_mapping[day]:
            # print(day)
            # print(hour)
            room_mapping[day][hour] = {}
        
        if classroom not in room_mapping[day][hour]:
            # print(classroom)
            room_mapping[day][hour][classroom] = []
        
        component_list = [component_code, component_class]
        # print(component_list)
        room_mapping[day][hour][classroom].append(component_list)

        # if component_code not in code_class_mapping:
        #     code_class_mapping[component_code] = []

        # if component_class not in code_class_mapping[component_code]:
        #     code_class_mapping[component_code].append(component_class)

    return room_mapping#, code_class_mapping

def get_students_mapping_from_pickle_file(datadir):
    dfpath = os.path.join(os.getcwd(), datadir, "dataframes.pickle")
    data = load_pickle_file(dfpath)

    students_mapping = {}
    students = {}

    for subject in data['data']:
        # print(subject)
        # input()
        component_code = subject[0][1][0].split('-')[0].strip() #0 - elemento dentro da lista, 1 - coluna com as informações, 0 - informações da disciplina, código é a primeira string antes do hífen
        component_class = subject[0][1][1].strip()  #0 - elemento dentro da lista, 1 - coluna com as informações, 0 - informações da turma
        students_ids = subject[1]['Matrícula'].tolist()
        students_courses = subject[1]['Curso'].tolist()

        if component_code not in students_mapping: #adiciona o código da disciplina apenas quando aparece pela primeira vez
            students_mapping[component_code] = {}
        
        if component_class not in students_mapping[component_code]: #os estudantes em uma turma de um determinado código de disciplina são os mesmos, então só adicionamos na primeira vez que aparecem
            students_mapping[component_code][component_class] = []

            for i in range(0, len(students_ids)):
                student_id = students_ids[i]
                student_course = students_courses[i]
                new_student = Student(student_id, student_course)

                if student_id not in students:
                    students[student_id] = new_student
                    students_mapping[component_code][component_class].append(new_student)
                else:
                    students_mapping[component_code][component_class].append(students[student_id])
    
    return students_mapping, students

def remove_data_not_imported(room_mapping, students_mapping):
    week_days = range(2, 7) #exclusivo
    class_hours = [8, 10, 14, 16, 18, 20]
    notfound = []
    sheet_component_code = []
    remove_dict = []
    sheet_mapping = {}

    for day in week_days:
        for hour in class_hours:
            # print(day, hour)
            # print(room_mapping[day][hour])
            # input()
            for classroom in room_mapping[day][hour]:
                for component_code, component_class in room_mapping[day][hour][classroom]:

                    if component_code not in sheet_mapping:
                        sheet_mapping[component_code] = []

                    if component_class not in sheet_mapping[component_code]:
                        sheet_mapping[component_code].append(component_class)
                        # print(component_code, component_class, sheet_mapping[component_code])
                        # input()
                        # print(students_mapping[component_code].keys())
                        # input()
    for component_code in sheet_mapping:
        if component_code not in students_mapping:
            remove_dict.append(component_code)
        else:
            students_mapping_classes = list(students_mapping[component_code].keys())
            for component_class in sheet_mapping[component_code]:
                if component_class not in students_mapping_classes:
                    print("NÃO ESTÁ: %s %s" % (component_code, component_class))
                    print(students_mapping_classes)
                else:
                    print("ESTÁ: %s %s" % (component_code, component_class))
                input()

    # print(sheet_mapping)
    # print("-----------")
                    # notfound_boolean = False
                    # sheet_component_code.append(component_code)
                    # if component_code not in students_mapping:
                    #     notfound_boolean = True    
                    # elif component_class == None:
                    #     if len(students_mapping[component_code]) > 1:
                    #         notfound_boolean = True
                    #     else:
                    #         #para pegar o código que tiver para o componente sem classe (talvez remover isso)
                    #         component_class = list(students_mapping[component_code].keys())[0]
                    # elif component_class not in students_mapping[component_code]:
                    #     notfound_boolean = True
                    # if notfound_boolean == True:
                    #     remove_dict.append([day, hour, classroom])
                    #     if component_code not in notfound:
                    #         notfound.append(component_code)
                    # print(day, hour, classroom, component_code, component_class)

    # for day, hour, classroom in remove_dict:
    #     del room_mapping[day][hour][classroom]

    # sheet_component_code = set(sheet_component_code)
    # sigaa_component_code = set(list(students_mapping.keys()))
    # notfound = set(notfound)
    # notimported_code = sheet_component_code - sigaa_component_code
    # notimported_class = list(set(notfound) - set(notimported_code))

    # print("CÓDIGOS DE DISCIPLINA NA PLANILHA QUE NÃO FORAM IMPORTADOS DO SIGAA: ", len(notimported_code))
    # print("CÓDIGOS DE TURMA NA PLANILHA QUE NÃO FORAM IMPORTADOS/ENCONTRADOS DO SIGAA: ", len(notimported_class))

def simulation(room_mapping, students_mapping, students):
    week_days = range(2, 7) #exclusivo
    class_hours = [8, 10, 14, 16, 18, 20]
    
    remove_data_not_imported(room_mapping, students_mapping)
    # input()

    #sábado e domingo
    for student_id in students:
        students[student_id].infectionProbability = random.uniform(0, 1) #editar depois
    
    #segunda a sexta
    for day in week_days:
        for hour in class_hours:
            for classroom in room_mapping[day][hour]:
                for component_code, component_class in room_mapping[day][hour][classroom]:
                    # if len(students_mapping[component_code]) > 1:
                    #     print(list(students_mapping[component_code].keys()))
                    #     print(component_code)
                    #     for component_class in students_mapping[component_code]:
                    #         print(component_class, len(students_mapping[component_code][component_class]))
                    #     # input()
                    if component_class == None:
                        if len(students_mapping[component_code]) == 1:    
                            component_class = list(students_mapping[component_code].keys())[0]
                    # print(day, hour, classroom, component_code, component_class)
                        
                    students_in_classroom = students_mapping[component_code][component_class]
                    #ADICIONAR FÓRMULA PARA O CÁLCULO DA PROBABILIDADE DE CONTAMINAÇÃO
                    
                    # for student in students_in_classroom:
                    #     print(student.id)

def load_file_parameters(filename):
    with open(filename, "r") as file:
        file_content = file.readlines()
        filepath = file_content[0].replace("\n", "")
        worksheet_name = file_content[1].replace("\n", "")
    
    return filepath, worksheet_name

def main():
    (opt, args) = parse_terminal_options()
    filepath, worksheet_name = load_file_parameters("file.txt")

    students_mapping, students = get_students_mapping_from_pickle_file(opt.datadir)
    room_mapping = get_room_mapping_from_excel_file(filepath, worksheet_name)
    # print(room_mapping)
    # simulation(room_mapping, students_mapping, students)
    remove_data_not_imported(room_mapping, students_mapping)

    return

if __name__ == "__main__":
    main()
