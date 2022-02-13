import os
import pandas as pd
import pickle
import requests
from bs4 import BeautifulSoup as BS
from getpass import getpass
from openpyxl import load_workbook
from optparse import OptionParser

dataframes = []

def parse_terminal_options():
    # a função gera padrões de parâmetros a serem analisados e retorna o resultado final

    parser = OptionParser()
    parser.add_option("-l", "--url", type="string", dest="url", default="https://sigaa.ufrrj.br", help="URL do Sigaa")
    parser.add_option("-d", "--datadir", type="string", dest="datadir", default="data", help="Diretório para salvar dados")
    parser.add_option("-v", "--verbose", action="store_true", dest="verbose", default=False, help="Imprime mensagens para debug")

    return parser.parse_args()

def insert_or_load_credentials(opt):
    # a função carrega as credenciais de um arquivo ou cria um arquivo com as credenciais informadas

    dirpath = os.path.join(os.getcwd(), opt.datadir)
    if not os.path.exists(dirpath):
        print("Criando o diretório de dados.")
        os.mkdir(dirpath)

    userinfo = os.path.join(os.getcwd(), opt.datadir, "userinfo.pickle")
    if os.path.exists(userinfo):
        print("Lendo credenciais de um arquivo.")
        file = open(userinfo, 'rb')
        data = pickle.load(file)
        file.close()
        opt.username = data['username']
        opt.password = data['password']
    else:
        print("Informe as suas credenciais do SIGAA.")
        opt.username = input("Usuário: ")
        opt.password = getpass("Senha: ")
        file = open(userinfo, 'wb')
        pickle.dump({'username': opt.username, "password": opt.password}, file)
        file.close()

def login_sigaa(opt, session, account):
    # a função faz login no sigaa e retorna um objeto html da página inicial

    data = {'user.login': opt.username,
            'user.senha': opt.password}
    
    # discente após o login é redirecionado para a página inicial do sigaa; docente apenas faz login
    html_homepage = session.post(url = opt.url + "/sigaa/logar.do?dispatch=logOn", data = data)

    if account == 'docente':
        html_homepage = session.get(url = opt.url + "/sigaa/verPortalDocente.do")

    return html_homepage

def get_opened_search_menu_html(opt, session, html_homepage, account):
    # a função abre o menu de busca de turmas do sigaa e retorna um objeto html da página de busca

    data = {
        'javax.faces.ViewState': get_view_state(html_homepage.content)
    }

    html_menuOpen = ''
    url_string = ''
    
    if account == 'discente':
        data['menu:form_menu_discente'] = 'menu:form_menu_discente'
        data['jscook_action'] = 'menu_form_menu_discente_j_id_jsp_1051466817_98_menu:A]#{ buscaTurmaBean.popularBuscaGeral }'
        data['id'] = '65996' #getId(html_homepage.content)
        url_string = "/sigaa/portais/discente/discente.jsf"

    elif account == 'docente':
        data['menu:form-menu_docente'] = 'menu:form-menu_docente'
        data['menu:ensCons_Turmas'] = 'menu:ensCons_Turmas'
        url_string = "/sigaa/portais/docente/docente.jsf"

    html_menuOpen = session.post(url = opt.url + url_string, data = data)

    return html_menuOpen

def search_and_save_class_infos(opt, session, html_menuOpen, account, year, semester):
    # a função faz a busca por turmas utilizando as informações carregadas do mapa de distribuição de salas e salva a resposta recebida em um dataframe

    room_mapping = get_room_mapping_from_excel_file("arquivo_turnos_unificados.xlsx", "Horários")

    data = {
        'form': 'form',
        'form:checkNivel': 'on',
        'form:selectNivelTurma': 'G',
        'form:checkAnoPeriodo': 'on',
        'form:inputAno': year,
        'form:inputPeriodo': semester,
        'form:selectUnidade': '0',
        'form:checkCodigo': 'on',
        'form:inputCodDisciplina': '',
        'form:checkCodigoTurma': '',
        'form:inputCodTurma': '',
        'form:selectCurso': '0',
        'form:selectSituacaoTurma': '1',
        'form:selectTipoTurma': '0',
        'form:selectModalidade': '0',
        'form:selectOpcaoOrdenacao': '1',
        'turmasEAD': 'false',
        'form:buttonBuscar': 'Buscar',
        'javax.faces.ViewState': get_view_state(html_menuOpen.content)
    }

    if account == 'docente':
        data['form:selectUnidade'] = '7579'
        data['form:inputLocal'] = ''
        data['form:inputHorario'] = ''
        data['form:inputNomeDisciplina'] = ''
        data['form:inputNomeDocente'] = ''
    elif account == 'discente':
        data['form:checkNivel'] = 'on'
        data['form:checkCodigoTurma'] = ''

    days = [2, 3, 4, 5, 6]
    hours = [8, 10, 14, 16, 20]

    for day in days:
        for hour in hours:
            for info in room_mapping[day][hour]:
                _, component_code, component_class = info 
                print("\n---------------------------------------\n")
                print(day, hour, component_code, component_class)

                data['form:inputCodDisciplina'] = component_code

                if component_class != None:
                    data['form:checkCodigoTurma'] = 'on'
                    data['form:inputCodTurma'] = component_class
                else:
                    data['form:checkCodigoTurma'] = ''
                    data['form:inputCodTurma'] = ''

                html_menuSearch = session.post(url = opt.url + "/sigaa/ensino/turma/busca_turma.jsf", data = data)
                save_infos_resulted_from_search(opt, session, html_menuSearch, data, account)

def save_infos_resulted_from_search(opt, session, html_menuSearch, data, account):
    # a função salva a resposta recebida sobre uma determinada turma pesquisada em um dataframe

    new_data = dict(data)
    new_data['form:buttonBuscar'] = ''
    new_data['id'] = get_id_turma(html_menuSearch.content, account)
    new_data['javax.faces.ViewState'] = get_view_state(html_menuSearch.content)

    if account == 'discente':
        new_data['form:sltbtnView'] = 'form:sltbtnView'
    elif account == 'docente':
        new_data['form:sltbtnSelecionarTurma'] = 'form:sltbtnSelecionarTurma'

    if new_data['id'] == -1:
        return -1
    
    html_menuResult = session.post(url = opt.url + "/sigaa/ensino/turma/busca_turma.jsf", data = new_data)
    print_html_tables(html_menuResult.content)
    save_html_tables(html_menuResult.content, account)

    return 0

def get_view_state(page_html):
    # a função busca e retorna o valor da variável ViewState na página html

    parsed_html = BS(page_html, 'html.parser')
    return parsed_html.find('input', attrs={'id' : 'javax.faces.ViewState'}).get('value')

def get_id_turma(page_html, account):
    # a função busca e retorna o valor inteiro do id atribuído uma turma na página html

    if account == 'discente':
        form_name = 'form:sltbtnTurmaVirtual'
        variable_name = 'idTurma'
    elif account == 'docente':
        form_name = 'form:sltbtnSelecionarTurma'
        variable_name = 'id'

    parsed_html = BS(page_html, 'html.parser')

    if parsed_html.find('a', attrs={'id' : form_name}) == None:
        return -1

    element = parsed_html.find('a', attrs={'id' : form_name}).get('onclick').split(',')
    element = [subelement for subelement in element if variable_name in subelement]
    
    return int(element[0].split(':')[1][1:-1])

def load_html_file(filepath):
    # a função carrega de um arquivo html o conteúdo dessa página

    file_t = open(filepath, "r")
    page_html = " ".join(file_t.readlines())
    file_t.close()
    return page_html

def get_room_mapping_from_excel_file(filepath, worksheet_name):
    # a função carrega dados de uma planilha; após isso, cria e retorna um dicionário da forma {day: {hour: [room, component_code, component_class] ...} ...}
    
    excel_workbook = load_workbook(filepath)
    worksheet = excel_workbook[worksheet_name]
    room_mapping = {}

    for i in range(2, worksheet.max_row + 1):
        day = worksheet.cell(row = i, column = 1).value
        hour = worksheet.cell(row = i, column = 2).value
        room = worksheet.cell(row = i, column = 3).value
        component_code = worksheet.cell(row = i, column = 4).value
        component_class = worksheet.cell(row = i, column = 5).value

        if day not in room_mapping:
            room_mapping[day] = {}
            room_mapping[day][hour] = []
        elif hour not in room_mapping[day]:
            room_mapping[day][hour] = []
        
        room_mapping[day][hour].append([room, component_code, component_class])
    return room_mapping

def print_html_tables(page_html, account):
    # a função imprime a lista de tabelas em uma página html, como uma lista de dataframes (fica mais agradável de visualizar); obs.: a primeira tabela normalmente é a mais importante

    dataframe_list = pd.read_html(page_html)
    if account == 'discente':
       print(dataframe_list[0])
    elif account == 'docente':
        print(dataframe_list)

def save_html_tables(page_html, account):
    # a função salva a lista de tabelas em uma página html em uma lista de dataframes

    dataframe_list = pd.read_html(page_html)
    if account == 'discente':
        dataframes.append(dataframe_list[0])
    elif account == 'docente':
        dataframes.append(dataframe_list)

def main():
    # carrega as opções escolhidas
    (opt, args) = parse_terminal_options()

    # carrega ou pede a inserção das credenciais que serão utilizadas
    insert_or_load_credentials(opt)

    # cria uma sessão
    session = requests.Session()

    # faz login e entra na página inicial
    html_homepage = login_sigaa(opt, session, 'discente')

    # abre a página do menu de busca de turmas
    html_menuOpen = get_opened_search_menu_html(opt, session, html_homepage, 'discente')

    # faz a busca por informações de diferentes disciplinas presentes em um arquivo externo e salva em uma lista de dataframes
    search_and_save_class_infos(opt, session, html_menuOpen, 'discente', 'ano escolhido', 'semestre escolhido')

    # salva a lista de dataframes em um arquivo pickle
    file = open(os.path.join(os.getcwd(), opt.datadir, "dataframes.pickle"), 'wb')
    pickle.dump({'data': dataframes}, file)
    file.close()
    
    return

if __name__ == "__main__":
    main()
