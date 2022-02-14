import os
from openpyxl import load_workbook

import classes
import file_operations

def get_room_mapping_from_excel_file(filepath, worksheet_name):
    # a função carrega dados de uma planilha; após isso, cria e retorna um dicionário da forma {day: {hour: [room, component_code, component_class] ...} ...}

    excel_workbook = load_workbook(filepath)
    worksheet = excel_workbook[worksheet_name]
    room_mapping = {}

    for i in range(2, worksheet.max_row + 1):
        day = worksheet.cell(row = i, column = 1).value
        hour = worksheet.cell(row = i, column = 2).value
        classroom = worksheet.cell(row = i, column = 3).value
        component_code = worksheet.cell(row = i, column = 4).value
        component_class = worksheet.cell(row = i, column = 5).value

        if day not in room_mapping:
            room_mapping[day] = {}
            room_mapping[day][hour] = {}
        elif hour not in room_mapping[day]:
            room_mapping[day][hour] = {}
        
        if classroom not in room_mapping[day][hour]:
            room_mapping[day][hour][classroom] = []
        
        component_list = [component_code, component_class]
        room_mapping[day][hour][classroom].append(component_list)

    return room_mapping

def get_academics_mapping_from_pickle_file(filepath):
    data = file_operations.load_pickle_file(filepath)

    academics_mapping = {}

    students = {} # usado como estrutura auxiliar
    professors_names = [] # usado para verificação apenas
    professors = []

    for infos in data['data']:
        infos_subject = infos[0] # duas colunas, contendo as informações da disciplina; primeira com os nomes das infos e segunda com: disciplina, turma, nome do docente
        infos_subject = infos_subject[1] # a segunda coluna possui os valores que estamos querendo utilizar

        component_code = infos_subject[0] # a primeira linha tem o código e o nome da disciplina
        component_code = component_code.split('-')[0].strip() # "código - nome" -> "código " -> "código"

        component_class = infos_subject[1].strip() # segunda linha tem o código da turma

        professor_name = infos_subject[2] # a terceira linha tem o nome do professor

        infos_students = infos[1] # informações dos alunos: matrícula, nome, curso

        students_ids = infos_students['Matrícula'].tolist() # lista com as matrículas
        students_courses = infos_students['Curso'].tolist() # lista com os cursos

        if component_code not in academics_mapping: # adiciona o código da disciplina apenas quando aparece pela primeira vez
            academics_mapping[component_code] = {}
        
        if component_class not in academics_mapping[component_code]: # os estudantes em uma turma de um determinado código de disciplina são os mesmos, então só adicionamos na primeira vez que aparecem
            academics_mapping[component_code][component_class] = []

            if professor_name not in professors_names:
                component_professor = classes.Professor(professor_name)
                professors_names.append(professor_name)
                professors.append(component_professor)
            else:
                index = professors_names.index(professor_name)
                component_professor = professors[index]

            academics_mapping[component_code][component_class].append(component_professor)

            for i in range(0, len(students_ids)):
                student_id = students_ids[i]
                student_course = students_courses[i]
                new_student = classes.Student(student_id, student_course)

                if student_id not in students:
                    students[student_id] = new_student
                    academics_mapping[component_code][component_class].append(new_student)
                else:
                    academics_mapping[component_code][component_class].append(students[student_id])            
    
    return academics_mapping

def remove_data_not_imported(room_mapping, academics_mapping):
    for day in room_mapping:
        for hour in room_mapping[day]:
            for classroom in room_mapping[day][hour]:
                indexes_to_remove = []
                count = 0
                for component_code, _ in room_mapping[day][hour][classroom]:
                    if component_code not in academics_mapping:
                        indexes_to_remove.append(count)
                    count += 1
                count = 0
                for remove_index in indexes_to_remove:
                    del room_mapping[day][hour][classroom][remove_index - count]
                    count += 1

def fill_empty_classes(room_mapping, academics_mapping):
    auxiliar = {}

    for component_code in academics_mapping:
        auxiliar[component_code] = list(academics_mapping[component_code].keys())

    for _ in range(0, 2):
        values_to_remove = []
        for day in room_mapping:
            for hour in room_mapping[day]:
                for classroom in room_mapping[day][hour]:
                    count = 0
                    tuples_to_modify = []
                    for component_code, component_class in room_mapping[day][hour][classroom]:
                        imported_classes = auxiliar[component_code]
                        classes_count = len(imported_classes)
                        if component_class == None:
                            if classes_count == 1:
                                tuples_to_modify.append((count, imported_classes[0]))
                                if (component_code, imported_classes[0]) not in values_to_remove:
                                    values_to_remove.append((component_code, imported_classes[0]))
                        else:
                            if component_class in auxiliar[component_code]:
                                if (component_code, component_class) not in values_to_remove:
                                    values_to_remove.append((component_code, component_class))
                        count += 1
                    
                    for modify_index, new_component_class in tuples_to_modify:
                        room_mapping[day][hour][classroom][modify_index][1] = new_component_class 
        for component_code, component_class in values_to_remove:
            auxiliar[component_code].remove(component_class)

if __name__ == "__main__":
    xlsx_filepath, worksheet_name, pickle_filepath = file_operations.load_parameters_from_txt_file("file.txt")
    
    room_mapping = get_room_mapping_from_excel_file(xlsx_filepath, worksheet_name)
    academics_mapping = get_academics_mapping_from_pickle_file(pickle_filepath)

    remove_data_not_imported(room_mapping, academics_mapping)
    fill_empty_classes(room_mapping, academics_mapping)
    
    file_operations.save_room_mapping_to_csv_file(room_mapping)
    file_operations.save_academics_mapping_to_csv_file(academics_mapping)
